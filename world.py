from bs4 import BeautifulSoup
import requests 
import openpyxl


wb = openpyxl.load_workbook('race_tracks.xlsx') 
wb.create_sheet('World_map')
print(wb.sheetnames)
sheets =wb['World_map']

url = requests.get('https://statisticstimes.com/geography/countries-by-continents.php').text
soup = BeautifulSoup(url,'lxml') 
countries_data = soup.find_all('tbody')
count = 0
country_list = list()
continent_list = list()
for data in countries_data:
    rows = data.find_all('tr')
    for row in rows:
        row_list = row.find_all('td', class_= 'name')
        count = count +1
        if count > 32:
            country = (row_list[0].text)
            country_list.append(country)
            country_continent = (row_list[3].text)
            continent_list.append(country_continent)
            print(f'{country} is in {country_continent}')

def merge_list(list1,list2):
    mereged_list = tuple(zip(list1,list2))
    return mereged_list

country_tuple = merge_list(country_list,continent_list)

rows = country_tuple
for row in rows:
    sheets.append(row)
excel = 'race_track_World_map.xlsx'
wb.save(excel)
print(f'\n Finished loading data into {excel}')
        
        
        
        
